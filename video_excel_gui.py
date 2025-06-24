import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import os
import sys
from datetime import datetime
import queue
import psutil
import win32gui
import win32process

# video_excel_processor 모듈 import
from video_excel_processor import VideoExcelProcessor

class VideoExcelGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("동영상/이미지 → 엑셀 처리기")
        self.root.geometry("800x600")
        
        # 처리 상태
        self.is_processing = False
        self.processor = None
        
        # 로그 큐 (스레드 간 통신)
        self.log_queue = queue.Queue()
        
        self.setup_ui()
        self.check_log_queue()
        
        # 윈도우 닫기 이벤트 처리
        self.root.protocol("WM_DELETE_WINDOW", self.exit_application)
        
    def setup_ui(self):
        """UI 구성"""
        # 메인 프레임
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 파일 선택 섹션
        file_frame = ttk.LabelFrame(main_frame, text="파일 및 폴더 선택", padding="10")
        file_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # Excel 파일 선택
        ttk.Label(file_frame, text="Excel 파일 (sample.xlsx):").grid(row=0, column=0, sticky=tk.W, pady=2)
        self.excel_path = tk.StringVar()
        ttk.Entry(file_frame, textvariable=self.excel_path, width=60).grid(row=0, column=1, padx=(10, 5), pady=2)
        ttk.Button(file_frame, text="찾기", command=self.select_excel_file).grid(row=0, column=2, pady=2)
        
        # 작업 폴더 선택 (입상관, 횡주관이 있는 상위 폴더)
        ttk.Label(file_frame, text="작업 폴더 (입상관/횡주관 포함):").grid(row=1, column=0, sticky=tk.W, pady=2)
        self.work_folder = tk.StringVar()
        ttk.Entry(file_frame, textvariable=self.work_folder, width=60).grid(row=1, column=1, padx=(10, 5), pady=2)
        ttk.Button(file_frame, text="찾기", command=self.select_work_folder).grid(row=1, column=2, pady=2)
        
        # 처리 버튼 섹션
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=1, column=0, columnspan=2, pady=10)
        
        self.start_button = ttk.Button(button_frame, text="시작", command=self.start_processing)
        self.start_button.pack(side=tk.LEFT, padx=5)
        
        self.stop_button = ttk.Button(button_frame, text="중지", command=self.stop_processing, state=tk.DISABLED)
        self.stop_button.pack(side=tk.LEFT, padx=5)
        
        self.exit_button = ttk.Button(button_frame, text="종료", command=self.exit_application)
        self.exit_button.pack(side=tk.LEFT, padx=5)
        
        # 진행률 바
        self.progress = ttk.Progressbar(main_frame, mode='indeterminate')
        self.progress.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        
        # 로그 섹션
        log_frame = ttk.LabelFrame(main_frame, text="처리 로그", padding="5")
        log_frame.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(10, 0))
        
        self.log_text = scrolledtext.ScrolledText(log_frame, height=20, width=80)
        self.log_text.pack(fill=tk.BOTH, expand=True)
        
        # 그리드 가중치 설정
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(3, weight=1)
        file_frame.columnconfigure(1, weight=1)
        
    def select_excel_file(self):
        """Excel 파일 선택"""
        filename = filedialog.askopenfilename(
            title="Excel 파일 선택",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.excel_path.set(filename)
            
    def select_work_folder(self):
        """작업 폴더 선택"""
        folder = filedialog.askdirectory(title="작업 폴더 선택 (입상관/횡주관 포함)")
        if folder:
            self.work_folder.set(folder)
            
    def log_message(self, message):
        """로그 메시지 추가"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_entry = f"[{timestamp}] {message}\n"
        self.log_queue.put(log_entry)
        
    def check_log_queue(self):
        """로그 큐 확인 및 UI 업데이트"""
        try:
            while True:
                log_entry = self.log_queue.get_nowait()
                self.log_text.insert(tk.END, log_entry)
                self.log_text.see(tk.END)
        except queue.Empty:
            pass
        
        # 100ms마다 다시 확인
        self.root.after(100, self.check_log_queue)
        
    def validate_inputs(self):
        """입력값 검증"""
        if not self.excel_path.get():
            messagebox.showerror("오류", "Excel 파일을 선택해주세요.")
            return False
            
        if not os.path.exists(self.excel_path.get()):
            messagebox.showerror("오류", "선택한 Excel 파일이 존재하지 않습니다.")
            return False
            
        if not self.work_folder.get():
            messagebox.showerror("오류", "작업 폴더를 선택해주세요.")
            return False
            
        if not os.path.exists(self.work_folder.get()):
            messagebox.showerror("오류", "선택한 작업 폴더가 존재하지 않습니다.")
            return False
            
        # 입상관, 횡주관 폴더 확인
        work_path = self.work_folder.get()
        if not (os.path.exists(os.path.join(work_path, "입상관")) or 
                os.path.exists(os.path.join(work_path, "횡주관"))):
            messagebox.showwarning("경고", "선택한 폴더에 '입상관' 또는 '횡주관' 폴더가 없습니다.\n계속 진행하시겠습니까?")
            
        return True
        
    def start_processing(self):
        """처리 시작"""
        if not self.validate_inputs():
            return
            
        # Excel 파일 사용 상태 확인
        if not self.check_excel_file_status():
            return
            
        self.is_processing = True
        self.start_button.config(state=tk.DISABLED)
        self.stop_button.config(state=tk.NORMAL)
        self.exit_button.config(state=tk.DISABLED)
        self.progress.start()
        
        # 로그 초기화
        self.log_text.delete(1.0, tk.END)
        self.log_message("처리를 시작합니다...")
        
        # 별도 스레드에서 처리 실행
        thread = threading.Thread(target=self.process_files, daemon=True)
        thread.start()
        
    def stop_processing(self):
        """처리 중지"""
        self.is_processing = False
        self.log_message("처리 중지 요청됨...")
        
    def is_excel_file_open(self, file_path):
        """Excel 파일이 열려있는지 확인"""
        try:
            filename = os.path.basename(file_path)
            
            # 프로세스 목록에서 Excel 프로세스 찾기
            for proc in psutil.process_iter(['pid', 'name']):
                try:
                    if proc.info['name'].lower() in ['excel.exe', 'et.exe']:  # MS Excel, WPS Office
                        # 윈도우 제목에서 파일명 확인
                        def enum_windows_callback(hwnd, windows):
                            if win32gui.IsWindowVisible(hwnd):
                                _, pid = win32process.GetWindowThreadProcessId(hwnd)
                                if pid == proc.info['pid']:
                                    title = win32gui.GetWindowText(hwnd)
                                    windows.append(title)
                            return True
                        
                        windows = []
                        win32gui.EnumWindows(enum_windows_callback, windows)
                        
                        for title in windows:
                            if filename.lower() in title.lower():
                                return True
                                
                except (psutil.NoSuchProcess, psutil.AccessDenied, Exception):
                    continue
                    
            return False
            
        except Exception as e:
            # 확인 실패 시 안전하게 False 반환
            return False
            
    def check_excel_file_status(self):
        """Excel 파일 상태 확인"""
        if not self.excel_path.get():
            return True
            
        if self.is_excel_file_open(self.excel_path.get()):
            result = messagebox.askyesno(
                "파일 사용 중", 
                f"선택한 Excel 파일이 현재 열려있습니다:\n{os.path.basename(self.excel_path.get())}\n\n"
                "파일이 열려있는 상태에서 처리하면 오류가 발생할 수 있습니다.\n"
                "Excel 파일을 닫고 다시 시도하시겠습니까?\n\n"
                "계속 진행하려면 '예'를, 취소하려면 '아니오'를 선택하세요."
            )
            return result
            
        return True
        
    def exit_application(self):
        """애플리케이션 종료"""
        if self.is_processing:
            result = messagebox.askyesno(
                "처리 중", 
                "현재 파일 처리가 진행 중입니다.\n"
                "정말로 종료하시겠습니까?"
            )
            if not result:
                return
                
            # 처리 중지
            self.is_processing = False
            self.log_message("애플리케이션 종료 중...")
            
        self.root.quit()
        self.root.destroy()
        
    def process_files(self):
        """파일 처리 (별도 스레드)"""
        try:
            # 작업 디렉토리 변경
            original_dir = os.getcwd()
            os.chdir(self.work_folder.get())
            
            # 처리기 생성
            excel_file = os.path.basename(self.excel_path.get())
            
            # Excel 파일을 작업 폴더로 복사 (필요한 경우)
            import shutil
            if not os.path.exists(excel_file):
                shutil.copy2(self.excel_path.get(), excel_file)
                self.log_message(f"Excel 파일 복사: {excel_file}")
            
            # 커스텀 처리기 생성
            processor = CustomVideoExcelProcessor(excel_file, None, None, self.log_message, self)
            processor.process_all()
            
            self.log_message("모든 처리가 완료되었습니다!")
            
        except Exception as e:
            self.log_message(f"처리 중 오류 발생: {str(e)}")
            
        finally:
            # 원래 디렉토리로 복원
            os.chdir(original_dir)
            
            # UI 상태 복원
            self.root.after(0, self.processing_finished)
            
    def processing_finished(self):
        """처리 완료 후 UI 상태 복원"""
        self.is_processing = False
        self.start_button.config(state=tk.NORMAL)
        self.stop_button.config(state=tk.DISABLED)
        self.exit_button.config(state=tk.NORMAL)
        self.progress.stop()


class CustomVideoExcelProcessor(VideoExcelProcessor):
    """GUI용 커스텀 처리기"""
    
    def __init__(self, excel_file, video_folder, image_folder, log_callback, gui_instance):
        super().__init__(excel_file, video_folder, image_folder)
        self.log_callback = log_callback
        self.gui = gui_instance
        self.processed_files = 0
        self.total_files = 0
        
    def log(self, message):
        """로그 출력"""
        if self.log_callback:
            self.log_callback(message)
            
    def count_total_files(self):
        """전체 파일 수 계산"""
        total = 0
        for folder_name in ["입상관", "횡주관"]:
            if os.path.exists(folder_name):
                for filename in os.listdir(folder_name):
                    if (filename.endswith('.mp4') or 
                        filename.lower().endswith(('.jpg', '.jpeg', '.png'))):
                        total += 1
        return total
        
    def process_all(self):
        """전체 처리 실행 (GUI용 오버라이드)"""
        if not self.load_excel():
            return
            
        # 전체 파일 수 계산
        self.total_files = self.count_total_files()
        self.log(f"처리할 파일 수: {self.total_files}개")
        
        try:
            # 입상관 폴더 처리
            if os.path.exists("입상관"):
                self.log("=== 입상관 파일 처리 시작 ===")
                self.process_folder("입상관", "입상")
            
            # 횡주관 폴더 처리
            if os.path.exists("횡주관"):
                self.log("=== 횡주관 파일 처리 시작 ===")
                self.process_folder("횡주관", "횡주")
            
            self.save_excel()
            
        finally:
            # 작업 완료 후 캡처 이미지 정리
            self.cleanup_captured_images()
            
    def process_folder(self, folder_path, pipe_type):
        """폴더 처리 (GUI용 오버라이드)"""
        if not os.path.exists(folder_path):
            self.log(f"폴더를 찾을 수 없습니다: {folder_path}")
            return
        
        # 캡처 이미지 저장할 폴더 생성
        capture_dir = os.path.join(os.getcwd(), 'captured_images')
        os.makedirs(capture_dir, exist_ok=True)
        
        # 파일 목록 가져오기
        files = [f for f in os.listdir(folder_path) 
                if (f.endswith('.mp4') or f.lower().endswith(('.jpg', '.jpeg', '.png')))]
        
        for filename in files:
            # 중지 요청 확인
            if hasattr(self.gui, 'is_processing') and not self.gui.is_processing:
                self.log("처리가 중지되었습니다.")
                return
                
            self.processed_files += 1
            progress_msg = f"[{self.processed_files}/{self.total_files}] {filename}"
            
            if filename.endswith('.mp4'):
                # 동영상 처리
                video_info = self.extract_video_info(filename, pipe_type)
                if not video_info:
                    self.log(f"❌ {progress_msg} - 파일명 패턴 불일치")
                    continue
                
                self.log(f"🎬 {progress_msg}")
                
                # 해당 단지, 유형 워크시트 선택
                worksheet = self.get_or_create_worksheet(video_info['complex'], pipe_type)
                
                # 해당하는 행 찾거나 생성
                if pipe_type == '횡주':
                    row = self.find_or_create_row(worksheet, pipe_type, video_info['dong'], 
                                                video_info['ho'], video_info['usage'], 
                                                video_info['line_detail'])
                else:
                    row = self.find_or_create_row(worksheet, pipe_type, video_info['dong'], 
                                                video_info['ho'], video_info['usage'])
                if not row:
                    continue
                
                # 동영상 캡처
                video_path = os.path.join(folder_path, filename)
                captured_files = self.capture_video_frames(video_path, capture_dir)
                
                if len(captured_files) >= 3:
                    # 이미지 삽입
                    self.insert_video_images(worksheet, pipe_type, captured_files, row)
                    self.log(f"✅ {filename} - 동영상 처리 완료")
                else:
                    self.log(f"❌ {filename} - 프레임 캡처 실패")
            
            elif filename.lower().endswith(('.jpg', '.jpeg', '.png')):
                # 이미지 처리
                image_info = self.extract_image_info(filename, pipe_type)
                if not image_info:
                    self.log(f"❌ {progress_msg} - 파일명 패턴 불일치")
                    continue
                
                self.log(f"🖼️ {progress_msg}")
                
                # 해당 단지, 유형 워크시트 선택
                worksheet = self.get_or_create_worksheet(image_info['complex'], pipe_type)
                
                # 해당하는 행 찾거나 생성
                if pipe_type == '횡주':
                    row = self.find_or_create_row(worksheet, pipe_type, image_info['dong'], 
                                                image_info['ho'], image_info['usage'], 
                                                image_info['line_detail'])
                else:
                    row = self.find_or_create_row(worksheet, pipe_type, image_info['dong'], 
                                                image_info['ho'], image_info['usage'])
                if not row:
                    continue
                
                # 이미지 및 텍스트 정보 입력
                self.process_issue_image(worksheet, folder_path, filename, image_info, row)
                self.log(f"✅ {filename} - 이미지 처리 완료")
                
    def insert_video_images(self, worksheet, pipe_type, captured_files, row):
        """동영상 이미지 삽입"""
        if pipe_type == '입상':
            position_col = self.find_column_by_name(worksheet, '위치사진')
            check1_col = self.find_column_by_name(worksheet, '점검사진1')
            check2_col = self.find_column_by_name(worksheet, '점검사진2')
            
            if position_col:
                self.insert_image_to_cell(worksheet, captured_files[0], row, position_col)
            if check1_col:
                self.insert_image_to_cell(worksheet, captured_files[1], row, check1_col)
            if check2_col:
                self.insert_image_to_cell(worksheet, captured_files[2], row, check2_col)
        
        else:  # 횡주
            position_col = self.find_column_by_name(worksheet, '위치사진')
            check1_col = 7
            check2_col = 8
            
            if position_col:
                self.insert_image_to_cell(worksheet, captured_files[0], row, position_col)
            self.insert_image_to_cell(worksheet, captured_files[1], row, check1_col)
            self.insert_image_to_cell(worksheet, captured_files[2], row, check2_col)
            
    def process_issue_image(self, worksheet, folder_path, filename, image_info, row):
        """이상 이미지 처리"""
        # 컬럼 번호 찾기
        issue_image_col = self.find_column_by_name(worksheet, '이상배관사진')
        issue_col = self.find_column_by_name(worksheet, '이상유무')
        location_col = self.find_column_by_name(worksheet, '위치')
        
        # 이미지 삽입
        if issue_image_col:
            image_path = os.path.join(folder_path, filename)
            self.insert_image_to_cell(worksheet, image_path, row, issue_image_col)
        
        # 텍스트 정보 입력
        if issue_col:
            worksheet.cell(row, issue_col).value = image_info['issue']
        if location_col:
            worksheet.cell(row, location_col).value = image_info['location']
            
    def insert_image_to_cell(self, worksheet, image_path, row, col):
        """이미지 삽입 (로그 제거)"""
        try:
            # 이미지 크기 조정
            resized_image = self.resize_image_for_excel(image_path)
            
            # 엑셀에 이미지 삽입
            from openpyxl.drawing.image import Image as OpenpyxlImage
            img = OpenpyxlImage(resized_image)
            
            # 셀 위치 계산
            cell_address = worksheet.cell(row, col).coordinate
            img.anchor = cell_address
            
            # 행 높이와 열 너비 조정
            worksheet.row_dimensions[row].height = 74
            worksheet.column_dimensions[worksheet.cell(row, col).column_letter].width = 13
            
            worksheet.add_image(img)
            return True
        except Exception as e:
            self.log(f"이미지 삽입 실패: {e}")
            return False


def main():
    root = tk.Tk()
    app = VideoExcelGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main() 