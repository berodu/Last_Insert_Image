#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
동영상 파일에서 캡처 이미지를 생성하고 엑셀 파일에 삽입하는 스크립트
"""

import cv2
import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image, ImageOps
import tempfile
from pathlib import Path
import hashlib
import shutil

class VideoExcelProcessor:
    def __init__(self, excel_file, video_folder, image_folder=None):
        self.excel_file = excel_file
        self.video_folder = video_folder
        self.image_folder = image_folder
        self.workbook = None
        self.worksheets = {}  # 단지별, 유형별 워크시트 저장 {(complex, type): worksheet}
    
    def get_complex_number(self, dong):
        """동 번호에서 단지 번호 추출"""
        dong_num = int(dong.replace('동', ''))
        return dong_num // 100  # 1101 -> 11, 901 -> 9, 301 -> 3 등
        
    def get_or_create_worksheet(self, complex_num, pipe_type):
        """필요한 워크시트를 가져오거나 생성"""
        key = (complex_num, pipe_type)
        
        # 이미 생성된 워크시트가 있으면 반환
        if key in self.worksheets:
            return self.worksheets[key]
        
        # 새 워크시트 생성
        sheet_name = f"점검결과사진({pipe_type})_{complex_num}단지"
        
        if sheet_name not in self.workbook.sheetnames:
            # 템플릿 복사해서 새 시트 생성
            template_sheet = self.workbook[f"{pipe_type}sample"]
            new_sheet = self.workbook.copy_worksheet(template_sheet)
            new_sheet.title = sheet_name
            print(f"새 시트 생성: {sheet_name}")
        else:
            print(f"기존 시트 사용: {sheet_name}")
        
        worksheet = self.workbook[sheet_name]
        self.worksheets[key] = worksheet
        return worksheet

    def load_excel(self):
        """엑셀 파일 로드 (시트는 필요할 때 동적 생성)"""
        try:
            self.workbook = load_workbook(self.excel_file)
            
            # 템플릿 시트 확인
            if "입상sample" not in self.workbook.sheetnames or "횡주sample" not in self.workbook.sheetnames:
                print("입상sample 또는 횡주sample 시트를 찾을 수 없습니다.")
                return False
            
            print(f"엑셀 파일 로드 완료: {self.excel_file}")
            return True
            
        except Exception as e:
            print(f"엑셀 파일 로드 실패: {e}")
            return False
    
    def extract_video_info(self, filename, pipe_type):
        """동영상 파일명에서 정보 추출"""
        # (이상배관) 부분 제거
        clean_filename = re.sub(r'\(이상배관\)', '', filename)
        
        if pipe_type == '입상':
            # 예: "1102동 4호 입상관 세탁.mp4"
            pattern = r'(\d+동)\s+(\d+호)\s+입상관\s+(.+)\.mp4'
            match = re.match(pattern, clean_filename)
            
            if match:
                dong = match.group(1)  # "1102동"
                ho = match.group(2)    # "4호"  
                usage = match.group(3) # "세탁"
                
                # 단지 구분
                complex_num = self.get_complex_number(dong)
                
                return {
                    'dong': dong,
                    'ho': ho,
                    'usage': usage,
                    'complex': complex_num,
                    'type': '입상'
                }
        
        elif pipe_type == '횡주':
            # 예: "1101동 1-1호 횡주관 배수.mp4"
            pattern = r'(\d+동)\s+(\d+-\d+호)\s+횡주관\s+(.+)\.mp4'
            match = re.match(pattern, clean_filename)
            
            if match:
                dong = match.group(1)      # "1101동"
                full_ho = match.group(2)   # "1-1호"
                usage = match.group(3)     # "배수"
                
                # 1-1호에서 1호와 1-1로 분리
                ho_pattern = r'(\d+)-(\d+)호'
                ho_match = re.match(ho_pattern, full_ho)
                if ho_match:
                    ho = f"{ho_match.group(1)}호"  # "1호"
                    line_detail = f"{ho_match.group(1)}-{ho_match.group(2)}"  # "1-1"
                else:
                    print(f"호수 패턴 불일치: {full_ho}")
                    return None
                
                # 단지 구분
                complex_num = self.get_complex_number(dong)
                
                return {
                    'dong': dong,
                    'ho': ho,
                    'line_detail': line_detail,
                    'usage': usage,
                    'complex': complex_num,
                    'type': '횡주'
                }
        
        print(f"파일명 패턴 불일치: {filename}")
        return None
    
    def extract_image_info(self, filename, pipe_type):
        """이미지 파일명에서 정보 추출"""
        if pipe_type == '입상':
            # 예: "1102동 4호 입상관 세탁_이물질_옥상.jpg"
            pattern = r'(\d+동)\s+(\d+호)\s+입상관\s+(.+?)_(.+?)_(.+?)\.(jpg|jpeg|png)'
            match = re.match(pattern, filename, re.IGNORECASE)
            
            if match:
                dong = match.group(1)    # "1102동"
                ho = match.group(2)      # "4호"
                usage = match.group(3)   # "세탁"
                issue = match.group(4)   # "이물질"
                location = match.group(5) # "옥상"
                
                # 단지 구분
                complex_num = self.get_complex_number(dong)
                
                return {
                    'dong': dong,
                    'ho': ho,
                    'usage': usage,
                    'issue': issue,
                    'location': location,
                    'complex': complex_num,
                    'type': '입상'
                }
        
        elif pipe_type == '횡주':
            # 예: "1101동 1-1호 횡주관 배수_이물질_옥상.jpg"
            pattern = r'(\d+동)\s+(\d+-\d+호)\s+횡주관\s+(.+?)_(.+?)_(.+?)\.(jpg|jpeg|png)'
            match = re.match(pattern, filename, re.IGNORECASE)
            
            if match:
                dong = match.group(1)      # "1101동"
                full_ho = match.group(2)   # "1-1호"
                usage = match.group(3)     # "배수"
                issue = match.group(4)     # "이물질"
                location = match.group(5)  # "옥상"
                
                # 1-1호에서 1호와 1-1로 분리
                ho_pattern = r'(\d+)-(\d+)호'
                ho_match = re.match(ho_pattern, full_ho)
                if ho_match:
                    ho = f"{ho_match.group(1)}호"  # "1호"
                    line_detail = f"{ho_match.group(1)}-{ho_match.group(2)}"  # "1-1"
                else:
                    print(f"호수 패턴 불일치: {full_ho}")
                    return None
                
                # 단지 구분
                complex_num = self.get_complex_number(dong)
                
                return {
                    'dong': dong,
                    'ho': ho,
                    'line_detail': line_detail,
                    'usage': usage,
                    'issue': issue,
                    'location': location,
                    'complex': complex_num,
                    'type': '횡주'
                }
        
        print(f"이미지 파일명 패턴 불일치: {filename}")
        return None
    
    def capture_video_frames(self, video_path, output_dir):
        """동영상에서 3개 프레임 캡처"""
        cap = cv2.VideoCapture(video_path)
        if not cap.isOpened():
            print(f"동영상 열기 실패: {video_path}")
            return []
        
        total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        fps = cap.get(cv2.CAP_PROP_FPS)
        duration = total_frames / fps
        
        # 캡처할 시간 계산
        times = [2.0, duration/2, max(2.0, duration-2.0)]
        captured_files = []
        
        # 파일명용 해시 생성
        file_hash = hashlib.md5(video_path.encode()).hexdigest()[:8]
        
        for i, time_sec in enumerate(times):
            frame_number = int(time_sec * fps)
            cap.set(cv2.CAP_PROP_POS_FRAMES, frame_number)
            
            ret, frame = cap.read()
            if ret:
                suffix = ['start', 'middle', 'end'][i]
                output_file = os.path.join(output_dir, f"capture_{file_hash}_{suffix}.jpg")
                
                # PIL을 사용해서 한글 경로 문제 해결
                try:
                    frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                    pil_image = Image.fromarray(frame_rgb)
                    pil_image.save(output_file, 'JPEG', quality=90)
                    captured_files.append(output_file)
                    print(f"캡처 완료: {output_file}")
                except Exception as e:
                    print(f"프레임 저장 실패: {e}")
            else:
                print(f"프레임 캡처 실패: {time_sec}초")
        
        cap.release()
        return captured_files
    
    def resize_image_for_excel(self, image_path, width=102, height=96):
        """엑셀에 삽입할 이미지 크기 조정"""
        try:
            with Image.open(image_path) as img:
                # 비율 무시하고 정확한 크기로 조정
                img_resized = img.resize((width, height), Image.Resampling.LANCZOS)
                
                # 임시 파일로 저장
                temp_file = tempfile.NamedTemporaryFile(suffix='.jpg', delete=False)
                img_resized.save(temp_file.name, 'JPEG', quality=70)
                return temp_file.name
        except Exception as e:
            print(f"이미지 크기 조정 실패: {e}")
            return image_path
    
    def find_column_by_name(self, worksheet, column_name):
        """컬럼명으로 컬럼 번호 찾기"""
        # 헤더 행은 3번째 행
        header_row = 3
        for col in range(1, worksheet.max_column + 1):
            cell_value = str(worksheet.cell(header_row, col).value or '').strip()
            if column_name == cell_value:
                return col
        return None

    def find_or_create_row(self, worksheet, pipe_type, dong, ho, usage, line_detail=None):
        """해당하는 행을 찾거나 새로 생성"""
        # 헤더 행은 3번째 행
        header_row = 3
        
        # 컬럼 위치 찾기
        dong_col = self.find_column_by_name(worksheet, '동')
        ho_col = self.find_column_by_name(worksheet, '라인')
        
        if pipe_type == '입상':
            usage_col = self.find_column_by_name(worksheet, '용도')
            pipe_col = self.find_column_by_name(worksheet, '배관경')
        else:  # 횡주
            usage_col = 4  # 횡주는 용도가 4번째 컬럼
            pipe_col = 5   # 횡주는 배관경이 5번째 컬럼
            line_detail_col = 3  # 횡주는 3번째 컬럼에 라인 상세 (1-1)
        
        if not all([dong_col, ho_col, usage_col, pipe_col]):
            print("필수 컬럼을 찾을 수 없습니다.")
            return None
        
        # 기존 행에서 매칭되는 행 찾기
        for row in range(header_row + 1, worksheet.max_row + 1):
            dong_value = str(worksheet.cell(row, dong_col).value or '').strip()
            ho_value = str(worksheet.cell(row, ho_col).value or '').strip()
            usage_value = str(worksheet.cell(row, usage_col).value or '').strip()
            
            if pipe_type == '횡주' and line_detail:
                line_detail_value = str(worksheet.cell(row, line_detail_col).value or '').strip()
                if (dong == dong_value and ho == ho_value and usage == usage_value 
                    and line_detail == line_detail_value):
                    print(f"기존 행 찾음: 행 {row}")
                    return row
            else:
                if dong == dong_value and ho == ho_value and usage == usage_value:
                    print(f"기존 행 찾음: 행 {row}")
                    return row
        
        # 새 행 생성 (빈 행 찾기)
        new_row = header_row + 1
        while worksheet.cell(new_row, dong_col).value:
            new_row += 1
        
        # 데이터 입력
        worksheet.cell(new_row, dong_col).value = dong
        worksheet.cell(new_row, ho_col).value = ho
        worksheet.cell(new_row, usage_col).value = usage
        worksheet.cell(new_row, pipe_col).value = "100A"  # 배관경 고정값
        
        if pipe_type == '횡주' and line_detail:
            worksheet.cell(new_row, line_detail_col).value = line_detail
            print(f"새 행 생성: 행 {new_row} - {dong} {ho} ({line_detail}) {usage}")
        else:
            print(f"새 행 생성: 행 {new_row} - {dong} {ho} {usage}")
        
        return new_row

    def insert_image_to_cell(self, worksheet, image_path, row, col):
        """엑셀 셀에 이미지 삽입"""
        try:
            # 이미지 크기 조정
            resized_image = self.resize_image_for_excel(image_path)
            
            # 엑셀에 이미지 삽입
            img = OpenpyxlImage(resized_image)
            
            # 셀 위치 계산
            cell_address = worksheet.cell(row, col).coordinate
            img.anchor = cell_address
            
            # 행 높이와 열 너비 조정 (이미지 크기에 맞춤)
            worksheet.row_dimensions[row].height = 74
            worksheet.column_dimensions[worksheet.cell(row, col).column_letter].width = 13
            
            worksheet.add_image(img)
            
            print(f"이미지 삽입 완료: {cell_address}")
            return True
        except Exception as e:
            print(f"이미지 삽입 실패: {e}")
            return False

    def process_folder(self, folder_path, pipe_type):
        """특정 폴더의 동영상과 이미지 처리"""
        if not os.path.exists(folder_path):
            print(f"폴더를 찾을 수 없습니다: {folder_path}")
            return
        
        print(f"\n=== {pipe_type} 파일 처리 중 ===")
        
        # 캡처 이미지 저장할 폴더 생성
        capture_dir = os.path.join(os.getcwd(), 'captured_images')
        os.makedirs(capture_dir, exist_ok=True)
        
        # 이미지 파일 그룹핑 (동, 호, 용도별로)
        image_groups = {}
        all_files = os.listdir(folder_path)
        
        # 이미지 파일들을 먼저 그룹핑
        for filename in all_files:
            if filename.lower().endswith(('.jpg', '.jpeg', '.png')):
                image_info = self.extract_image_info(filename, pipe_type)
                if image_info:
                    if pipe_type == '횡주':
                        key = (image_info['dong'], image_info['ho'], image_info['usage'], image_info['line_detail'])
                    else:
                        key = (image_info['dong'], image_info['ho'], image_info['usage'])
                    
                    if key not in image_groups:
                        image_groups[key] = []
                    image_groups[key].append((filename, image_info))
        
        # 동영상 파일 처리
        for filename in all_files:
            if filename.endswith('.mp4'):
                video_info = self.extract_video_info(filename, pipe_type)
                if not video_info:
                    continue
                
                print(f"동영상 처리 중: {filename}")
                
                # 해당 단지, 유형 워크시트 선택
                worksheet = self.get_or_create_worksheet(video_info['complex'], pipe_type)
                
                # 해당하는 행 찾거나 생성
                if pipe_type == '횡주':
                    row = self.find_or_create_row(worksheet, pipe_type, video_info['dong'], 
                                                video_info['ho'], video_info['usage'], 
                                                video_info['line_detail'])
                else:
                    row = self.find_or_create_row(worksheet, pipe_type, video_info['dong'], 
                                                video_info['ho'], video_info['usage'])
                if not row:
                    continue
                
                # 동영상 캡처
                video_path = os.path.join(folder_path, filename)
                captured_files = self.capture_video_frames(video_path, capture_dir)
                
                if len(captured_files) >= 3:
                    # 컬럼 번호 찾기
                    if pipe_type == '입상':
                        position_col = self.find_column_by_name(worksheet, '위치사진')
                        check1_col = self.find_column_by_name(worksheet, '점검사진1')
                        check2_col = self.find_column_by_name(worksheet, '점검사진2')
                        
                        # 이미지를 엑셀에 삽입
                        if position_col:
                            self.insert_image_to_cell(worksheet, captured_files[0], row, position_col)
                        if check1_col:
                            self.insert_image_to_cell(worksheet, captured_files[1], row, check1_col)
                        if check2_col:
                            self.insert_image_to_cell(worksheet, captured_files[2], row, check2_col)
                    
                    else:  # 횡주
                        position_col = self.find_column_by_name(worksheet, '위치사진')
                        # 횡주는 점검사진이 2개 (컬럼 7, 8)
                        check1_col = 7
                        check2_col = 8
                        
                        # 이미지를 엑셀에 삽입
                        if position_col:
                            self.insert_image_to_cell(worksheet, captured_files[0], row, position_col)
                        self.insert_image_to_cell(worksheet, captured_files[1], row, check1_col)
                        self.insert_image_to_cell(worksheet, captured_files[2], row, check2_col)
        
        # 이미지 파일 처리 (그룹별로 첫 번째만)
        processed_groups = set()
        for key, files_info in image_groups.items():
            if key in processed_groups:
                continue
            
            # 첫 번째 파일만 처리
            filename, image_info = files_info[0]
            total_count = len(files_info)
            
            print(f"이미지 처리 중: {filename} (총 {total_count}개 중 첫 번째)")
            
            # 해당 단지, 유형 워크시트 선택
            worksheet = self.get_or_create_worksheet(image_info['complex'], pipe_type)
            
            # 해당하는 행 찾거나 생성
            if pipe_type == '횡주':
                row = self.find_or_create_row(worksheet, pipe_type, image_info['dong'], 
                                            image_info['ho'], image_info['usage'], 
                                            image_info['line_detail'])
            else:
                row = self.find_or_create_row(worksheet, pipe_type, image_info['dong'], 
                                            image_info['ho'], image_info['usage'])
            if not row:
                continue
            
            # 컬럼 번호 찾기
            issue_image_col = self.find_column_by_name(worksheet, '이상배관사진')
            issue_col = self.find_column_by_name(worksheet, '이상유무')
            location_col = self.find_column_by_name(worksheet, '위치')
            
            # 이미지 삽입
            if issue_image_col:
                image_path = os.path.join(folder_path, filename)
                self.insert_image_to_cell(worksheet, image_path, row, issue_image_col)
            
            # 텍스트 정보 입력
            if issue_col:
                worksheet.cell(row, issue_col).value = image_info['issue']
            if location_col:
                # 위치 정보에 총 개수 추가
                location_text = f"{image_info['location']}({total_count})" if total_count > 1 else image_info['location']
                worksheet.cell(row, location_col).value = location_text
            
            processed_groups.add(key)

    def save_excel(self, output_file=None):
        """엑셀 파일 저장"""
        if not self.workbook:
            print("저장할 워크북이 없습니다.")
            return
        
        if not output_file:
            output_file = self.excel_file.replace('.xlsx', '_processed.xlsx')
        
        try:
            self.workbook.save(output_file)
            print(f"엑셀 파일 저장 완료: {output_file}")
        except Exception as e:
            print(f"엑셀 파일 저장 실패: {e}")

    def cleanup_captured_images(self):
        """캡처된 이미지 파일들 정리"""
        capture_dir = os.path.join(os.getcwd(), 'captured_images')
        if os.path.exists(capture_dir):
            try:
                shutil.rmtree(capture_dir)
                print(f"캡처 이미지 폴더 정리 완료: {capture_dir}")
            except Exception as e:
                print(f"캡처 이미지 정리 실패: {e}")

    def process_all(self):
        """전체 처리 실행"""
        if not self.load_excel():
            return
        
        try:
            # 입상관 폴더 처리
            if os.path.exists("입상관"):
                self.process_folder("입상관", "입상")
            
            # 횡주관 폴더 처리
            if os.path.exists("횡주관"):
                self.process_folder("횡주관", "횡주")
            
            self.save_excel()
            
        finally:
            # 작업 완료 후 캡처 이미지 정리
            self.cleanup_captured_images()

def main():
    # 파일 경로 설정
    excel_file = "sample.xlsx"
    
    print("=== 동영상/이미지 → 엑셀 처리 시작 ===")
    
    # 처리 실행
    processor = VideoExcelProcessor(excel_file, None, None)
    processor.process_all()
    
    print("=== 처리 완료 ===")

if __name__ == "__main__":
    main() 